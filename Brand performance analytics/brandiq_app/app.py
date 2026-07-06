import os
import re
import json
import uuid
from datetime import datetime, timedelta, date
from io import BytesIO

import pandas as pd
import numpy as np
from flask import Flask, render_template, request, jsonify, session, send_file
from flask_cors import CORS
from sklearn.linear_model import LinearRegression
import cohere
from dotenv import load_dotenv

from models import db, Brand, Product, Upload, Bill, BillItem, ReorderHistory, Prediction

app = Flask(__name__)
app.secret_key = 'brandiq-secret-key-2026'
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///brandiq.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024  # 50MB

# Ensure upload directory exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

load_dotenv()
COHERE_API_KEY = os.environ.get("COHERE_API_KEY", "")
co = cohere.Client(COHERE_API_KEY)

CORS(app)
db.init_app(app)

with app.app_context():
    db.create_all()

# ─── HELPERS ───────────────────────────────────────────────────────────────

def allowed_file(filename):
    return filename.lower().endswith(('.xlsx', '.xls', '.csv'))

def parse_date(val):
    """Parse date from various formats."""
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, pd.Timestamp):
        return val.date()
    if isinstance(val, str):
        val = val.strip()
        # Handle "2026-06-20 00:00:00" format
        if ' ' in val:
            val = val.split(' ')[0]
        for fmt in ['%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%Y/%m/%d']:
            try:
                return datetime.strptime(val, fmt).date()
            except:
                pass
    return None

def safe_float(val):
    if val is None:
        return 0.0
    try:
        return float(str(val).replace(',', '').strip())
    except:
        return 0.0

def safe_int(val):
    if val is None:
        return 0
    try:
        return int(float(str(val).replace(',', '').strip()))
    except:
        return 0

# ─── INVENTORY PARSER ──────────────────────────────────────────────────────

def parse_inventory(filepath):
    """Parse Inventory.xlsx and store brands + products."""
    df = pd.read_excel(filepath, sheet_name='Data', dtype=str)
    df = df.fillna('')
    
    count = 0
    for _, row in df.iterrows():
        barcode = str(row.get('Barcode', '')).strip()
        if not barcode:
            continue
        
        brand_name = str(row.get('Category 1', '')).strip()
        if not brand_name:
            continue
        
        # Upsert brand
        brand = Brand.query.filter_by(name=brand_name).first()
        if not brand:
            brand = Brand(name=brand_name, first_seen_at=datetime.utcnow())
            db.session.add(brand)
            db.session.commit()
        
        # Upsert product
        product = Product.query.filter_by(barcode=barcode).first()
        if not product:
            product = Product(barcode=barcode)
        
        product.item_code = str(row.get('Item Code ', '')).strip()
        product.item_name = str(row.get('Item Name', '')).strip()
        product.article_name = str(row.get('Article Name', '')).strip()
        product.brand_id = brand.id
        product.division = str(row.get('Division', '')).strip()
        product.section = str(row.get('Section ', '')).strip()
        product.department = str(row.get('Department', '')).strip()
        product.category2 = str(row.get('Category 2', '')).strip()
        product.color = str(row.get('Category 3', '')).strip()
        product.size = str(row.get('Category 4', '')).strip()
        product.material = str(row.get('Category 5', '')).strip()
        product.gender = str(row.get('Category 6', '')).strip()
        product.mrp = safe_float(row.get('MRP', 0))
        product.vendor_name = str(row.get('Vendor Name', '')).strip()
        
        db.session.add(product)
        count += 1
    
    db.session.commit()
    return count

# ─── BRAND EXTRACTION ──────────────────────────────────────────────────────

def extract_brand_from_item_name(item_name):
    """Extract brand name from item name string.
    
    Examples:
      '23SCO136 SainSisters 23-SCO-136 Pink OS Mul Women' -> 'SainSisters'
      '8903273000024 FAVORI 2223534132 Multicolor OS Cotton Unisex' -> 'FAVORI'
      'EKT-FG-BAG-0104 EKATRA EKT-FG-BAG-0104-ASSORTED Multicolor OS paper Unisex' -> 'EKATRA'
      'JG_027_ATT Jyotirgamaya JG_027_ATT Multicolor OS Essential oils Unisex' -> 'Jyotirgamaya'
    """
    if not item_name:
        return None
    
    # Split by spaces and look for the second word (after barcode)
    parts = item_name.strip().split()
    if len(parts) >= 2:
        # The second word is typically the brand name
        brand_candidate = parts[1]
        # Skip if it looks like a code/number
        if re.match(r'^[A-Za-z][A-Za-z\s&-]+$', brand_candidate) and len(brand_candidate) > 1:
            return brand_candidate
        # Also try third word if second looks like a code
        if len(parts) >= 3 and re.match(r'^[A-Za-z][A-Za-z\s&-]+$', parts[2]) and len(parts[2]) > 1:
            return parts[2]
    
    return None

def get_or_create_brand(brand_name):
    """Get existing brand or create a new one."""
    if not brand_name:
        return None
    brand = Brand.query.filter_by(name=brand_name).first()
    if not brand:
        brand = Brand(name=brand_name, first_seen_at=datetime.utcnow())
        db.session.add(brand)
        db.session.flush()
    return brand

def get_or_create_product(barcode, item_name, brand_id):
    """Get existing product or create a new one."""
    if not barcode:
        return None
    product = Product.query.filter_by(barcode=barcode).first()
    if not product:
        product = Product(barcode=barcode)
        product.item_name = item_name
        product.brand_id = brand_id
        db.session.add(product)
        db.session.flush()
    return product

# ─── POS BILL PARSER ───────────────────────────────────────────────────────

def parse_pos_bills(filepath, upload_id):
    """Parse POS Bill Register file (CSV or Excel) with 3-section nested format.
    
    Data structure per bill:
      Row 0 (header): [Store, Date, Date, Date, BillNo, GST Doc, Time, Cashier, Cashier, Customer, ...]
      Row 1: empty or sub-header
      Row 2: column headers for items (Barcode, Item Name, ...)
      Rows 3..N: items [Barcode, Item Name, ..., Qty, ..., Price, Basic, ..., Discount, ..., Net, ...]
      Then MOP row: [MOP, Type, ...] (optional, not in CSV format)
      Then MOP data: [PaymentMode, Type, ..., Tender, Balance, Collection, ...]
    """
    try:
        # Read file based on extension
        if filepath.lower().endswith('.csv'):
            df = pd.read_csv(filepath, dtype=str, header=None, keep_default_na=False)
        else:
            df = pd.read_excel(filepath, dtype=str, header=None)
        df = df.fillna('')
    except Exception as e:
        print(f"Error reading file: {e}")
        raise Exception(f"Failed to read file: {str(e)}")
    
    bills_parsed = 0
    items_parsed = 0
    
    i = 0
    while i < len(df):
        row = df.iloc[i]
        vals = [str(v).strip() for v in row.values]
        
        # Skip header/meta rows (first 5-6 rows are metadata)
        if vals[0] in ['ORCHID ENT OUTFIT PRIVATE LIMITED', '#POS Bill Register#', 'Owner Site', '']:
            if vals[0] == '' and i > 0:
                # Check previous row - if it was "Owner Site", this could be empty separator
                prev = df.iloc[i-1] if i > 0 else None
                if prev is not None and str(prev.values[0]).strip() not in ['Owner Site', '#POS Bill Register#', 'ORCHID ENT OUTFIT PRIVATE LIMITED', 'Show Item Details:', 'Period:']:
                    pass  # Might be a real row
                else:
                    i += 1
                    continue
            elif vals[0] in ['ORCHID ENT OUTFIT PRIVATE LIMITED', '#POS Bill Register#']:
                i += 1
                continue
            elif vals[0] == 'Owner Site':
                i += 1
                continue
            elif vals[0] == '' and i < 6:
                i += 1
                continue
        
        # Skip rows that start with 'Show Item', 'Period:', 'Print Date:'
        if vals[0].startswith('Show Item') or vals[0].startswith('Period:') or vals[0].startswith('Print Date:'):
            i += 1
            continue
        
        # Detect bill header: has a bill number pattern like TST00002/06-26 in column 4
        is_bill_header = False
        bill_no = ''
        bill_date = None
        customer_name = ''
        cashier = ''
        payment_mode = ''
        net_amount = 0
        discount_amount = 0
        total_amount = 0
        item_count = 0
        
        # Check if this row has a valid bill no in column 4
        if len(vals) > 4 and vals[4]:
            if re.match(r'^[A-Z0-9]+/\d{2}-\d{2}$', vals[4]) or re.match(r'^[A-Z]+\d+/\d{2}-\d{2}$', vals[4]):
                is_bill_header = True
        
        if not is_bill_header and len(vals) > 0:
            # Check if first column is a barcode (item row) - but skip if we're not in a bill
            if vals[0] not in ['MOP', 'Barcode', ''] and not vals[0].startswith('Toscee'):
                # Could be an orphan item row, skip
                i += 1
                continue
            i += 1
            continue
        
        if not is_bill_header:
            i += 1
            continue
        
        # Extract bill header data
        # Column layout: 0=OwnerSite, 1=Date, 2=Date, 3=Date, 4=BillNo, 5=GSTDoc, 6=Time, 7=Cashier, 8=Cashier, 9=Customer, 10=Prints, 11=Items, 12=Items, 13=SaleQty, 14=ReturnQty, 15=SaleAmt, 16=ReturnAmt, 17=BasicAmt, 18=Discount, 19=Net, 20=RoundOff, 21=NetPayable
        bill_no = vals[4]
        # Try multiple date column positions (1, 2, or 3) to handle different export formats
        bill_date = None
        for date_col in [1, 2, 3]:
            if len(vals) > date_col and vals[date_col]:
                bill_date = parse_date(vals[date_col])
                if bill_date:
                    break
        customer_name = vals[9] if len(vals) > 9 else ''
        cashier = vals[7] if len(vals) > 7 else ''
        total_amount = safe_float(vals[15]) if len(vals) > 15 else 0  # Sale Amount
        discount_amount = safe_float(vals[18]) if len(vals) > 18 else 0  # Discount
        net_amount = safe_float(vals[19]) if len(vals) > 19 else 0  # Net Amount
        item_count = safe_int(vals[11]) if len(vals) > 11 else 0
        
        if not bill_date:
            i += 1
            continue
        
        # Check for duplicate bill (same bill_no + same date)
        existing = Bill.query.filter_by(bill_no=bill_no, bill_date=bill_date).first()
        if existing:
            bills_parsed += 1
            # Skip to next bill header
            j = i + 1
            while j < len(df):
                check_row = df.iloc[j]
                check_vals = [str(v).strip() for v in check_row.values]
                if check_vals[0] == 'MOP':
                    j += 2
                    break
                if len(check_vals) > 4 and check_vals[4] and re.match(r'^[A-Z0-9]+/\d{2}-\d{2}$', check_vals[4]):
                    break
                j += 1
            i = j
            continue
        
        # Create the bill
        bill = Bill(
            bill_no=bill_no,
            bill_date=bill_date,
            customer_name=customer_name,
            cashier=cashier,
            total_amount=total_amount,
            discount_amount=discount_amount,
            net_amount=net_amount,
            item_count=item_count,
            upload_id=upload_id
        )
        db.session.add(bill)
        db.session.flush()
        
        # Move past header - skip empty/sub-header rows
        j = i + 1
        while j < len(df):
            check_row = df.iloc[j]
            check_vals = [str(v).strip() for v in check_row.values]
            
            # Skip empty rows and column header rows
            if check_vals[0] == '' or check_vals[0] == 'Barcode':
                j += 1
                continue
            
            # Check if we hit MOP section (MOP in column 0) or a new bill header
            if check_vals[0] == 'MOP':
                # Extract payment mode from MOP data row (next row or same row)
                if j + 1 < len(df):
                    mop_data = df.iloc[j + 1]
                    mop_data_vals = [str(v).strip() for v in mop_data.values]
                    # Payment mode is in column 0 of MOP data row
                    payment_mode = mop_data_vals[0] if len(mop_data_vals) > 0 else ''
                    if payment_mode in ['MOP', '', 'Type']:
                        payment_mode = ''
                bill.payment_mode = payment_mode
                db.session.add(bill)
                j += 2  # Skip MOP header + data rows
                break
            
            # Check if this is the start of a new bill
            if len(check_vals) > 4 and check_vals[4] and re.match(r'^[A-Z0-9]+/\d{2}-\d{2}$', check_vals[4]):
                break
            
            # Check if this is a print date/end row
            if check_vals[0].startswith('Print Date:') or check_vals[0].startswith('Page No.'):
                j += 1
                break
            
            # This is an item row - parse it
            barcode = check_vals[0]
            if not barcode or barcode in ['', 'Barcode']:
                j += 1
                continue
            
            item_name = check_vals[1] if len(check_vals) > 1 else ''
            qty = safe_int(check_vals[6]) if len(check_vals) > 6 else 1
            price = safe_float(check_vals[8]) if len(check_vals) > 8 else 0
            basic_amt = safe_float(check_vals[9]) if len(check_vals) > 9 else 0
            disc_amt = safe_float(check_vals[10]) if len(check_vals) > 10 else 0
            net_amt = safe_float(check_vals[12]) if len(check_vals) > 12 else 0
            
            # Find product by barcode (from inventory or auto-create)
            product = Product.query.filter_by(barcode=barcode).first()
            if not product:
                # Auto-extract brand and create product on the fly
                brand_name = extract_brand_from_item_name(item_name)
                brand = get_or_create_brand(brand_name) if brand_name else None
                product = get_or_create_product(barcode, item_name, brand.id if brand else None)
            
            bill_item = BillItem(
                bill_id=bill.id,
                barcode=barcode,
                product_id=product.id if product else None,
                item_name=item_name,
                quantity=qty,
                price=price,
                basic_amount=basic_amt,
                discount_amount=disc_amt,
                net_amount=net_amt
            )
            db.session.add(bill_item)
            items_parsed += 1
            j += 1
        
        bills_parsed += 1
        i = j  # Move to next section
    
    db.session.commit()
    return bills_parsed, items_parsed

# ─── ANALYTICS ENGINE ──────────────────────────────────────────────────────

def get_weekly_data(week_start=None, week_end=None):
    """Get aggregated data for a date range."""
    query = db.session.query(Bill)
    
    if week_start:
        query = query.filter(Bill.bill_date >= week_start)
    if week_end:
        query = query.filter(Bill.bill_date <= week_end)
    
    bills = query.all()
    
    total_revenue = sum(b.net_amount for b in bills)
    total_bills = len(bills)
    total_items = sum(b.item_count for b in bills)
    total_discount = sum(b.discount_amount for b in bills)
    avg_bill_value = total_revenue / total_bills if total_bills > 0 else 0
    
    return {
        'total_revenue': round(total_revenue, 2),
        'total_bills': total_bills,
        'total_items': total_items,
        'total_discount': round(total_discount, 2),
        'avg_bill_value': round(avg_bill_value, 2),
        'bills': bills
    }

def get_brand_performance(week_start=None, week_end=None):
    """Get brand-wise performance metrics."""
    query = db.session.query(BillItem).join(Bill).join(Product, BillItem.product_id == Product.id).join(Brand)
    
    if week_start:
        query = query.filter(Bill.bill_date >= week_start)
    if week_end:
        query = query.filter(Bill.bill_date <= week_end)
    
    results = query.with_entities(
        Brand.name,
        db.func.sum(BillItem.net_amount).label('revenue'),
        db.func.sum(BillItem.quantity).label('units'),
        db.func.count(db.distinct(Bill.id)).label('bills'),
        db.func.sum(BillItem.discount_amount).label('discount'),
        db.func.avg(BillItem.price).label('avg_price')
    ).group_by(Brand.name).all()
    
    total_revenue = sum(r.revenue or 0 for r in results)
    
    brands = []
    for r in results:
        share = (r.revenue / total_revenue * 100) if total_revenue > 0 else 0
        brands.append({
            'name': r.name,
            'revenue': round(r.revenue or 0, 2),
            'units': r.units or 0,
            'bills': r.bills or 0,
            'discount': round(r.discount or 0, 2),
            'avg_price': round(r.avg_price or 0, 2),
            'share': round(share, 2)
        })
    
    brands.sort(key=lambda x: x['revenue'], reverse=True)
    for idx, b in enumerate(brands):
        b['rank'] = idx + 1
    
    return brands

def get_previous_week_range(week_start):
    """Get the previous week's date range."""
    prev_start = week_start - timedelta(days=7)
    prev_end = week_start - timedelta(days=1)
    return prev_start, prev_end

def calculate_growth(current_brands, prev_brands):
    """Calculate week-over-week growth for each brand."""
    prev_map = {b['name']: b['revenue'] for b in prev_brands}
    
    for brand in current_brands:
        prev_rev = prev_map.get(brand['name'], 0)
        if prev_rev > 0:
            growth = ((brand['revenue'] - prev_rev) / prev_rev) * 100
        else:
            growth = 100 if brand['revenue'] > 0 else 0
        brand['growth'] = round(growth, 2)
        
        if growth > 10:
            brand['status'] = 'growing'
        elif growth < -10:
            brand['status'] = 'declining'
        else:
            brand['status'] = 'stable'
    
    return current_brands

def get_daily_trend(week_start, week_end):
    """Get daily revenue trend."""
    bills = Bill.query.filter(
        Bill.bill_date >= week_start,
        Bill.bill_date <= week_end
    ).order_by(Bill.bill_date).all()
    
    daily = {}
    for b in bills:
        d = str(b.bill_date)
        if d not in daily:
            daily[d] = {'date': d, 'revenue': 0, 'bills': 0, 'items': 0}
        daily[d]['revenue'] += b.net_amount
        daily[d]['bills'] += 1
        daily[d]['items'] += b.item_count
    
    return list(daily.values())

def get_payment_distribution(week_start, week_end):
    """Get payment mode distribution."""
    query = db.session.query(
        Bill.payment_mode,
        db.func.sum(Bill.net_amount).label('amount'),
        db.func.count(Bill.id).label('count')
    )
    
    if week_start:
        query = query.filter(Bill.bill_date >= week_start)
    if week_end:
        query = query.filter(Bill.bill_date <= week_end)
    
    results = query.group_by(Bill.payment_mode).all()
    
    return [{
        'mode': r.payment_mode or 'Unknown',
        'amount': round(r.amount or 0, 2),
        'count': r.count or 0
    } for r in results]

def get_top_products(week_start, week_end, limit=20):
    """Get top selling products."""
    query = db.session.query(
        BillItem.barcode,
        BillItem.item_name,
        Brand.name.label('brand_name'),
        db.func.sum(BillItem.quantity).label('units'),
        db.func.sum(BillItem.net_amount).label('revenue'),
        db.func.count(db.distinct(Bill.id)).label('bills')
    ).join(Bill).outerjoin(Product, BillItem.product_id == Product.id).outerjoin(Brand, Product.brand_id == Brand.id)
    
    if week_start:
        query = query.filter(Bill.bill_date >= week_start)
    if week_end:
        query = query.filter(Bill.bill_date <= week_end)
    
    results = query.group_by(BillItem.barcode, BillItem.item_name, Brand.name)\
                   .order_by(db.func.sum(BillItem.net_amount).desc())\
                   .limit(limit).all()
    
    return [{
        'barcode': r.barcode or '',
        'item_name': r.item_name or '',
        'brand': r.brand_name or 'Unknown',
        'units': r.units or 0,
        'revenue': round(r.revenue or 0, 2),
        'bills': r.bills or 0
    } for r in results]

def get_department_performance(week_start, week_end):
    """Get department-wise performance from inventory data."""
    query = db.session.query(
        Product.department,
        db.func.sum(BillItem.net_amount).label('revenue'),
        db.func.sum(BillItem.quantity).label('units'),
        db.func.count(db.distinct(Bill.id)).label('bills')
    ).join(BillItem, Product.id == BillItem.product_id)\
     .join(Bill)\
     .filter(Product.department != '')
    
    if week_start:
        query = query.filter(Bill.bill_date >= week_start)
    if week_end:
        query = query.filter(Bill.bill_date <= week_end)
    
    results = query.group_by(Product.department).all()
    
    return [{
        'department': r.department,
        'revenue': round(r.revenue or 0, 2),
        'units': r.units or 0,
        'bills': r.bills or 0
    } for r in results]

# ─── PREDICTION ENGINE ─────────────────────────────────────────────────────

def predict_next_week():
    """Predict next week's revenue based on historical weekly data."""
    # Get all bills grouped by week
    bills = Bill.query.order_by(Bill.bill_date).all()
    if len(bills) < 2:
        return None
    
    # Group by ISO week
    weekly = {}
    for b in bills:
        iso = b.bill_date.isocalendar()
        week_key = f"{iso[0]}-W{iso[1]:02d}"
        if week_key not in weekly:
            weekly[week_key] = {'revenue': 0, 'bills': 0, 'items': 0, 'start_date': b.bill_date - timedelta(days=b.bill_date.weekday())}
        weekly[week_key]['revenue'] += b.net_amount
        weekly[week_key]['bills'] += 1
        weekly[week_key]['items'] += b.item_count
    
    weeks = sorted(weekly.keys())
    if len(weeks) < 2:
        return None
    
    # Prepare data for linear regression
    X = np.array(range(len(weeks))).reshape(-1, 1)
    y = np.array([weekly[w]['revenue'] for w in weeks])
    
    model = LinearRegression()
    model.fit(X, y)
    
    next_week_idx = len(weeks)
    predicted_revenue = model.predict([[next_week_idx]])[0]
    
    # Calculate trend
    last_week_rev = weekly[weeks[-1]]['revenue']
    growth_rate = ((predicted_revenue - last_week_rev) / last_week_rev * 100) if last_week_rev > 0 else 0
    
    # Get last week's start date and calculate next week
    last_start = weekly[weeks[-1]]['start_date']
    next_start = last_start + timedelta(days=7)
    next_end = next_start + timedelta(days=6)
    
    return {
        'predicted_revenue': round(predicted_revenue, 2),
        'current_revenue': round(last_week_rev, 2),
        'growth_rate': round(growth_rate, 2),
        'next_week_start': str(next_start),
        'next_week_end': str(next_end),
        'confidence': round(model.score(X, y) * 100, 1),
        'trend': 'up' if growth_rate > 0 else 'down',
        'weekly_data': [
            {'week': w, 'revenue': round(weekly[w]['revenue'], 2), 'bills': weekly[w]['bills']}
            for w in weeks
        ]
    }

def predict_brand_trends():
    """Predict which brands are growing/declining."""
    # Get all brand-week data
    results = db.session.query(
        Brand.name,
        Bill.bill_date,
        db.func.sum(BillItem.net_amount).label('revenue')
    ).join(Product, Brand.id == Product.brand_id)\
     .join(BillItem, Product.id == BillItem.product_id)\
     .join(Bill)\
     .group_by(Brand.name, Bill.bill_date)\
     .order_by(Brand.name, Bill.bill_date).all()
    
    if not results:
        return []
    
    # Group by brand
    brand_data = {}
    for r in results:
        if r.name not in brand_data:
            brand_data[r.name] = []
        brand_data[r.name].append({
            'date': str(r.bill_date),
            'revenue': r.revenue or 0
        })
    
    trends = []
    for brand_name, data in brand_data.items():
        if len(data) < 2:
            continue
        
        # Simple linear trend
        X = np.array(range(len(data))).reshape(-1, 1)
        y = np.array([d['revenue'] for d in data])
        
        model = LinearRegression()
        model.fit(X, y)
        
        slope = model.coef_[0]
        avg_rev = np.mean(y)
        trend_pct = (slope / avg_rev * 100) if avg_rev > 0 else 0
        
        trends.append({
            'brand': brand_name,
            'slope': round(slope, 2),
            'trend_pct': round(trend_pct, 2),
            'avg_revenue': round(avg_rev, 2),
            'direction': 'up' if slope > 0 else 'down',
            'data_points': len(data)
        })
    
    trends.sort(key=lambda x: x['trend_pct'], reverse=True)
    return trends

def get_brand_growth_analytics(weeks=8):
    """Get detailed brand growth analytics for the last N weeks."""
    # Get the date range
    today = date.today()
    end_date = today - timedelta(days=today.weekday())  # Last Monday
    start_date = end_date - timedelta(weeks=weeks)  # N weeks ago
    
    # Get all brand-week combinations
    results = db.session.query(
        Brand.name,
        Bill.bill_date,
        db.func.sum(BillItem.net_amount).label('revenue'),
        db.func.sum(BillItem.quantity).label('units'),
        db.func.count(db.distinct(Bill.id)).label('bills')
    ).join(Product, Brand.id == Product.brand_id)\
     .join(BillItem, Product.id == BillItem.product_id)\
     .join(Bill)\
     .filter(Bill.bill_date >= start_date, Bill.bill_date <= end_date)\
     .group_by(Brand.name, Bill.bill_date)\
     .order_by(Brand.name, Bill.bill_date).all()
    
    # Group by brand
    brand_weekly = {}
    for r in results:
        if r.name not in brand_weekly:
            brand_weekly[r.name] = []
        brand_weekly[r.name].append({
            'date': str(r.bill_date),
            'revenue': round(r.revenue or 0, 2),
            'units': r.units or 0,
            'bills': r.bills or 0
        })
    
    # Calculate growth metrics for each brand
    brand_growth = []
    for brand_name, weekly_data in brand_weekly.items():
        if len(weekly_data) < 2:
            continue
        
        # Sort by date
        weekly_data.sort(key=lambda x: x['date'])
        
        # Calculate week-over-week growth
        growth_rates = []
        for i in range(1, len(weekly_data)):
            prev_rev = weekly_data[i-1]['revenue']
            curr_rev = weekly_data[i]['revenue']
            if prev_rev > 0:
                growth = ((curr_rev - prev_rev) / prev_rev) * 100
            else:
                growth = 100 if curr_rev > 0 else 0
            growth_rates.append(round(growth, 2))
        
        # Calculate metrics
        latest_revenue = weekly_data[-1]['revenue']
        previous_revenue = weekly_data[-2]['revenue']
        latest_growth = growth_rates[-1] if growth_rates else 0
        
        # Moving average (last 4 weeks)
        ma_4 = round(np.mean(growth_rates[-4:]), 2) if len(growth_rates) >= 4 else latest_growth
        
        # Consistency score (0-100) - lower std dev = higher consistency
        if len(growth_rates) > 1:
            std_dev = np.std(growth_rates)
            max_std = 50  # Normalize
            consistency = max(0, 100 - (std_dev / max_std * 100))
        else:
            consistency = 50
        consistency = round(consistency, 1)
        
        # Growth momentum (comparing last 2 weeks to previous 2 weeks)
        if len(growth_rates) >= 4:
            recent_avg = np.mean(growth_rates[-2:])
            older_avg = np.mean(growth_rates[-4:-2])
            momentum = round(recent_avg - older_avg, 2)
        else:
            momentum = 0
        
        # Growth streak
        streak = 0
        streak_direction = 'neutral'
        for g in reversed(growth_rates):
            if g > 0 and streak_direction in ['neutral', 'growing']:
                streak += 1
                streak_direction = 'growing'
            elif g < 0 and streak_direction in ['neutral', 'declining']:
                streak += 1
                streak_direction = 'declining'
            elif g == 0:
                break
            else:
                break
        
        # Total revenue over period
        total_revenue = sum(w['revenue'] for w in weekly_data)
        
        brand_growth.append({
            'brand': brand_name,
            'latest_revenue': round(latest_revenue, 2),
            'total_revenue': round(total_revenue, 2),
            'latest_growth': round(latest_growth, 2),
            'ma_4_weeks': round(ma_4, 2),
            'consistency': consistency,
            'momentum': round(momentum, 2),
            'streak': streak,
            'streak_direction': streak_direction,
            'data_points': len(weekly_data),
            'weekly_data': weekly_data,
            'growth_rates': growth_rates
        })
    
    # Sort by latest growth
    brand_growth.sort(key=lambda x: x['latest_growth'], reverse=True)
    
    return {
        'brands': brand_growth,
        'period': {
            'start': str(start_date),
            'end': str(end_date),
            'weeks': weeks
        }
    }

# ─── REORDER ENGINE ─────────────────────────────────────────────────────────

def get_reorder_suggestions():
    """ML-based reorder suggestions using sales velocity and trends."""
    # Pre-load inventory data for margin lookup (read once, not per product)
    inv_lookup = {}
    try:
        inv_path = os.path.join(os.path.dirname(__file__), '..', 'Inventory.xlsx')
        if os.path.exists(inv_path):
            inv_data = pd.read_excel(inv_path, sheet_name='Data', dtype=str)
            # Drop rows with missing barcode
            inv_data = inv_data.dropna(subset=['Barcode'])
            inv_data['Barcode'] = inv_data['Barcode'].astype(str).str.strip()
            inv_lookup = dict(zip(inv_data['Barcode'], inv_data['Description 3'].fillna('25').astype(str).str.strip()))
        else:
            print(f"Inventory file not found at: {os.path.abspath(inv_path)}")
    except Exception as e:
        import traceback
        print(f"Error reading inventory file: {e}")
        print(traceback.format_exc())
        
    products = Product.query.all()
    suggestions = []
    
    for product in products:
        if not product.barcode:
            continue
        
        barcode = product.barcode.strip()
        
        # Get sales history for this product (last 12 weeks)
        end_date = date.today()
        start_date = end_date - timedelta(weeks=12)
        
        sales_data = db.session.query(
            Bill.bill_date,
            db.func.sum(BillItem.quantity).label('qty')
        ).join(BillItem, Bill.id == BillItem.bill_id)\
         .filter(BillItem.product_id == product.id)\
         .filter(Bill.bill_date >= start_date)\
         .group_by(Bill.bill_date)\
         .order_by(Bill.bill_date).all()
        
        # Get brand name
        brand = Brand.query.get(product.brand_id) if product.brand_id else None
        brand_name = brand.name if brand else 'Unknown'
        
        if not sales_data:
            continue
        
        # Calculate weekly sales
        weekly_sales = {}
        for s in sales_data:
            week_start = s.bill_date - timedelta(days=s.bill_date.weekday())
            week_key = str(week_start)
            weekly_sales[week_key] = weekly_sales.get(week_key, 0) + (s.qty or 0)
        weeks_with_data = len(weekly_sales)
        
        total_qty = sum(weekly_sales.values())
        avg_weekly = total_qty / max(weeks_with_data, 1)
        safety_weeks = 4
        
        if weeks_with_data < 1:
            continue
        
        # Trend analysis
        sorted_weeks = sorted(weekly_sales.keys())
        if len(sorted_weeks) >= 2:
            X = np.array(range(len(sorted_weeks))).reshape(-1, 1)
            y = np.array([weekly_sales[w] for w in sorted_weeks])
            model = LinearRegression()
            model.fit(X, y)
            trend_slope = model.coef_[0]
            trend_pct = (trend_slope / max(avg_weekly, 0.01)) * 100
        else:
            trend_slope = 0
            trend_pct = 0
            
        if trend_pct > 10:
            priority = 'High'
        elif trend_pct < -10:
            priority = 'Low'
        else:
            priority = 'Medium'
        
        reorder_qty = max(1, round(avg_weekly * safety_weeks))
        
        # Get margin from pre-loaded inventory lookup
        margin = 25
        if barcode in inv_lookup:
            desc3 = inv_lookup[barcode]
            if desc3 and str(desc3).strip():
                try:
                    margin = float(str(desc3).strip())
                except:
                    pass
        
        suggestions.append({
            'barcode': barcode,
            'product_name': product.item_name or '',
            'brand': brand_name,
            'avg_weekly_sales': round(avg_weekly, 1),
            'trend_pct': round(trend_pct, 1),
            'trend_direction': 'up' if trend_pct > 0 else 'down',
            'priority': priority,
            'reorder_qty': reorder_qty,
            'margin': margin,
            'mrp': product.mrp or 0,
            'total_sold': total_qty,
            'weeks_data': weeks_with_data
        })
    
    priority_order = {'High': 0, 'Medium': 1, 'Low': 2}
    suggestions.sort(key=lambda x: (priority_order.get(x['priority'], 3), -x['reorder_qty']))
    
    return suggestions

def generate_po_for_brand(brand_name, suggestions, output_path):
    """Generate PO Excel file for a specific brand using the PT template."""
    import openpyxl
    from openpyxl.utils import get_column_letter
    
    # Resolve template path
    base_dir = os.path.join(os.path.dirname(__file__), '..')
    template_path = os.path.join(base_dir, 'PT file Template for ginesys.xlsx')
    
    # Load product details from database (Inventory was uploaded via UI)
    brand_obj = Brand.query.filter_by(name=brand_name).first()
    
    # Try to load template, or create a new workbook
    try:
        if os.path.exists(template_path):
            wb = openpyxl.load_workbook(template_path)
            ws = wb['Sheet2']
            # Clear existing data (keep header row 1)
            for row in range(ws.max_row, 1, -1):
                ws.delete_rows(row)
        else:
            raise FileNotFoundError("Template not found")
    except Exception as e:
        print(f"Warning: Template not found, creating new workbook: {e}")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet2"
        # Write header row
        headers = ['Brand', 'Category', 'Sub Category', 'Product', 'Article', 'Style Code',
                   'Color', 'Size', 'Barcode', 'MRP', 'HSN', 'GST%', 'Material', 'Gender',
                   'Season', 'Qty', 'PO Rate', 'UOM', 'Margin%', 'Net Price']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
    
    # Fill data row by row
    row_num = 2
    for s in suggestions:
        # Get product details from database
        product_db = Product.query.filter_by(barcode=s['barcode']).first()
        
        if product_db:
            category = product_db.division or ''
            sub_category = product_db.section or ''
            product = product_db.department or ''
            article = product_db.article_name or ''
            style_code = product_db.item_code or ''
            color = product_db.color or ''
            size = product_db.size or ''
            mrp = product_db.mrp or s['mrp']
            hsn = ''
            material = product_db.material or ''
            gender = product_db.gender or ''
            season = 'SS26'
            uom = 'PCS'
            margin = s['margin']
        else:
            category = ''
            sub_category = ''
            product = ''
            article = ''
            style_code = ''
            color = ''
            size = ''
            mrp = s['mrp']
            hsn = ''
            material = ''
            gender = ''
            season = 'SS26'
            uom = 'PCS'
            margin = s['margin']
        
        # Write data
        ws.cell(row=row_num, column=1, value=brand_name)
        ws.cell(row=row_num, column=2, value=category)
        ws.cell(row=row_num, column=3, value=sub_category)
        ws.cell(row=row_num, column=4, value=product)
        ws.cell(row=row_num, column=5, value=article)
        ws.cell(row=row_num, column=6, value=style_code)
        ws.cell(row=row_num, column=7, value=color)
        ws.cell(row=row_num, column=8, value=size)
        ws.cell(row=row_num, column=9, value=s['barcode'])
        ws.cell(row=row_num, column=10, value=mrp)
        ws.cell(row=row_num, column=11, value=hsn)
        ws.cell(row=row_num, column=12, value=0.05)  # GST% = 5%
        ws.cell(row=row_num, column=13, value=material)
        ws.cell(row=row_num, column=14, value=gender)
        ws.cell(row=row_num, column=15, value=season)
        ws.cell(row=row_num, column=16, value=s['reorder_qty'])
        # PO RATE formula: =ROUND((J/(1+L/100))*(1-S/100),2)
        ws.cell(row=row_num, column=17).value = f'=ROUND((J{row_num}/(1+L{row_num}/100))*(1-S{row_num}/100),2)'
        ws.cell(row=row_num, column=18, value=uom)
        ws.cell(row=row_num, column=19, value=margin)
        # Net Price formula: =J/(1+L)
        ws.cell(row=row_num, column=20).value = f'=J{row_num}/(1+L{row_num})'
        
        row_num += 1
    
    wb.save(output_path)
    return row_num - 2  # Number of items written

# ─── ROUTES ─────────────────────────────────────────────────────────────────

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/upload/inventory', methods=['POST'])
def upload_inventory():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if not file or not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file format. Use .xlsx or .csv'}), 400
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"inventory_{uuid.uuid4().hex}_{file.filename}")
    file.save(filepath)
    
    try:
        count = parse_inventory(filepath)
        
        upload = Upload(
            file_name=file.filename,
            file_type='inventory',
            total_rows=count,
            processed_rows=count,
            status='completed'
        )
        db.session.add(upload)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Inventory loaded: {count} products',
            'products_count': count
        })
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f"Error parsing inventory file: {error_details}")
        return jsonify({'error': f'Failed to parse: {str(e)}'}), 500

@app.route('/api/upload/pos', methods=['POST'])
def upload_pos():
    if 'file' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['file']
    if not file or not allowed_file(file.filename):
        return jsonify({'error': 'Invalid file format. Use .xlsx or .csv'}), 400
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], f"pos_{uuid.uuid4().hex}_{file.filename}")
    file.save(filepath)
    
    try:
        upload = Upload(
            file_name=file.filename,
            file_type='pos',
            status='processing'
        )
        db.session.add(upload)
        db.session.flush()
        
        bills_count, items_count = parse_pos_bills(filepath, upload.id)
        
        upload.status = 'completed'
        upload.total_rows = bills_count
        upload.processed_rows = items_count
        db.session.add(upload)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': f'Parsed {bills_count} bills with {items_count} items',
            'bills': bills_count,
            'items': items_count
        })
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        print(f"Error parsing POS file: {error_details}")
        return jsonify({'error': f'Failed to parse: {str(e)}'}), 500

@app.route('/api/dashboard', methods=['GET'])
def dashboard():
    week_start_str = request.args.get('week_start')
    week_end_str = request.args.get('week_end')
    
    # Default to current week if not specified
    if not week_start_str:
        today = date.today()
        week_start = today - timedelta(days=today.weekday())
    else:
        week_start = parse_date(week_start_str)
    
    if not week_end_str:
        week_end = week_start + timedelta(days=6)
    else:
        week_end = parse_date(week_end_str)
    
    # Get current week data
    current_data = get_weekly_data(week_start, week_end)
    
    # Get previous week data for comparison
    prev_start, prev_end = get_previous_week_range(week_start)
    prev_data = get_weekly_data(prev_start, prev_end)
    
    # Calculate growth
    revenue_growth = 0
    if prev_data['total_revenue'] > 0:
        revenue_growth = ((current_data['total_revenue'] - prev_data['total_revenue']) / prev_data['total_revenue']) * 100
    
    # Get brand performance
    current_brands = get_brand_performance(week_start, week_end)
    prev_brands = get_brand_performance(prev_start, prev_end)
    brands_with_growth = calculate_growth(current_brands, prev_brands)
    
    # Get other analytics
    daily_trend = get_daily_trend(week_start, week_end)
    payment_dist = get_payment_distribution(week_start, week_end)
    top_products = get_top_products(week_start, week_end)
    dept_perf = get_department_performance(week_start, week_end)
    
    # Get predictions
    prediction = predict_next_week()
    brand_trends = predict_brand_trends()
    
    return jsonify({
        'kpis': {
            'total_revenue': current_data['total_revenue'],
            'total_bills': current_data['total_bills'],
            'total_items': current_data['total_items'],
            'total_discount': current_data['total_discount'],
            'avg_bill_value': current_data['avg_bill_value'],
            'revenue_growth': round(revenue_growth, 2),
            'prev_revenue': prev_data['total_revenue'],
            'prev_bills': prev_data['total_bills']
        },
        'brands': brands_with_growth,
        'daily_trend': daily_trend,
        'payment_distribution': payment_dist,
        'top_products': top_products,
        'departments': dept_perf,
        'prediction': prediction,
        'brand_trends': brand_trends,
        'period': {
            'start': str(week_start),
            'end': str(week_end)
        }
    })

@app.route('/api/brands', methods=['GET'])
def brands():
    week_start_str = request.args.get('week_start')
    week_end_str = request.args.get('week_end')
    
    today = date.today()
    week_start = parse_date(week_start_str) if week_start_str else (today - timedelta(days=today.weekday()))
    week_end = parse_date(week_end_str) if week_end_str else (week_start + timedelta(days=6))
    
    prev_start, prev_end = get_previous_week_range(week_start)
    
    current_brands = get_brand_performance(week_start, week_end)
    prev_brands = get_brand_performance(prev_start, prev_end)
    brands_with_growth = calculate_growth(current_brands, prev_brands)
    
    return jsonify({'brands': brands_with_growth})

@app.route('/api/dates', methods=['GET'])
def get_dates():
    """Get the available date range from bills data."""
    first = Bill.query.order_by(Bill.bill_date).first()
    last = Bill.query.order_by(Bill.bill_date.desc()).first()
    
    if first and last:
        start = first.bill_date
        end = last.bill_date
        
        # Find Monday of the week containing the first bill
        start_monday = start - timedelta(days=start.weekday())
        # Default week end = that Monday + 6 days (Sunday of the same week)
        default_end = start_monday + timedelta(days=6)
        
        return jsonify({
            'min_date': str(start),
            'max_date': str(end),
            'default_week_start': str(start_monday),
            'default_week_end': str(default_end)
        })
    
    return jsonify({
        'min_date': None,
        'max_date': None,
        'default_week_start': None,
        'default_week_end': None
    })

@app.route('/api/brand/<int:brand_id>', methods=['GET'])
def brand_detail(brand_id):
    brand = Brand.query.get(brand_id)
    if not brand:
        return jsonify({'error': 'Brand not found'}), 404
    
    week_start_str = request.args.get('week_start')
    today = date.today()
    week_start = parse_date(week_start_str) if week_start_str else (today - timedelta(days=today.weekday()))
    week_end = week_start + timedelta(days=6)
    
    # Get brand products
    products = Product.query.filter_by(brand_id=brand_id).all()
    product_ids = [p.id for p in products]
    
    # Get brand sales
    items = db.session.query(BillItem).filter(
        BillItem.product_id.in_(product_ids),
        Bill.bill_date >= week_start,
        Bill.bill_date <= week_end
    ).join(Bill).all()
    
    total_revenue = sum(i.net_amount for i in items)
    total_units = sum(i.quantity for i in items)
    total_discount = sum(i.discount_amount for i in items)
    
    # Get weekly trend for this brand
    trend_data = db.session.query(
        Bill.bill_date,
        db.func.sum(BillItem.net_amount).label('revenue'),
        db.func.sum(BillItem.quantity).label('units')
    ).filter(
        BillItem.product_id.in_(product_ids),
        Bill.bill_date >= week_start - timedelta(days=90)
    ).join(Bill).group_by(Bill.bill_date).order_by(Bill.bill_date).all()
    
    return jsonify({
        'brand': {
            'id': brand.id,
            'name': brand.name,
            'products_count': len(products),
            'revenue': round(total_revenue, 2),
            'units': total_units,
            'discount': round(total_discount, 2)
        },
        'trend': [{'date': str(r.bill_date), 'revenue': round(r.revenue or 0, 2), 'units': r.units or 0} for r in trend_data]
    })

@app.route('/api/products', methods=['GET'])
def products():
    week_start_str = request.args.get('week_start')
    week_end_str = request.args.get('week_end')
    
    today = date.today()
    week_start = parse_date(week_start_str) if week_start_str else (today - timedelta(days=today.weekday()))
    week_end = parse_date(week_end_str) if week_end_str else (week_start + timedelta(days=6))
    
    top = get_top_products(week_start, week_end, limit=50)
    return jsonify({'products': top})

@app.route('/api/predictions', methods=['GET'])
def predictions():
    prediction = predict_next_week()
    brand_trends = predict_brand_trends()
    
    return jsonify({
        'revenue_prediction': prediction,
        'brand_trends': brand_trends
    })

@app.route('/api/brand-growth', methods=['GET'])
def brand_growth():
    weeks = request.args.get('weeks', 8, type=int)
    data = get_brand_growth_analytics(weeks=weeks)
    return jsonify(data)

# ─── CATEGORY/SUBCATEGORY/PRODUCT TRENDS ─────────────────────────────────────

@app.route('/api/trends', methods=['GET'])
def trends_analysis():
    """Get category, subcategory, and product wise trends."""
    weeks = request.args.get('weeks', 12, type=int)
    level = request.args.get('level', 'category')  # category, subcategory, product
    parent = request.args.get('parent', '')  # filter by parent category/subcategory
    start_date_str = request.args.get('start_date')
    end_date_str = request.args.get('end_date')
    
    if start_date_str and end_date_str:
        start_date = datetime.strptime(start_date_str, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date_str, '%Y-%m-%d').date()
        # Calculate prev_start based on the duration
        duration_days = (end_date - start_date).days
        prev_start = start_date - timedelta(days=duration_days)
    else:
        end_date = date.today()
        start_date = end_date - timedelta(weeks=weeks)
        prev_start = start_date - timedelta(weeks=weeks//2)
    
    # Build query based on level
    if level == 'category':
        group_col = Product.division
        name_col = Product.division
        label = 'category'
    elif level == 'subcategory':
        group_col = Product.section
        name_col = Product.section
        label = 'subcategory'
        # Filter by parent category if specified
        if parent:
            parent_filter = Product.division == parent
        else:
            parent_filter = True
    elif level == 'product':
        group_col = Product.department
        name_col = Product.department
        label = 'product'
        # Filter by parent subcategory if specified
        if parent:
            parent_filter = Product.section == parent
        else:
            parent_filter = True
    else:
        return jsonify({'error': 'Invalid level'}), 400
    
    # Get current period data
    current_query = db.session.query(
        group_col.label('name'),
        db.func.sum(BillItem.net_amount).label('revenue'),
        db.func.sum(BillItem.quantity).label('units'),
        db.func.count(db.distinct(Bill.id)).label('bills'),
        db.func.sum(BillItem.discount_amount).label('discount')
    ).join(Bill, Bill.id == BillItem.bill_id)\
     .join(Product, Product.id == BillItem.product_id)\
     .filter(Bill.bill_date >= start_date, Bill.bill_date <= end_date)
    
    if level != 'category' and parent:
        current_query = current_query.filter(parent_filter)
    
    # Filter out empty names
    current_results = current_query.filter(group_col != '', group_col.isnot(None))\
        .group_by(group_col).order_by(db.func.sum(BillItem.net_amount).desc()).all()
    
    # Get previous period data for growth calculation
    prev_query = db.session.query(
        group_col.label('name'),
        db.func.sum(BillItem.net_amount).label('revenue')
    ).join(Bill, Bill.id == BillItem.bill_id)\
     .join(Product, Product.id == BillItem.product_id)\
     .filter(Bill.bill_date >= prev_start, Bill.bill_date < start_date)
    
    if level != 'category' and parent:
        prev_query = prev_query.filter(parent_filter)
    
    prev_results = prev_query.filter(group_col != '', group_col.isnot(None))\
        .group_by(group_col).all()
    
    prev_map = {r.name: r.revenue or 0 for r in prev_results}
    
    # Get weekly trend data
    weekly_data = db.session.query(
        group_col.label('name'),
        Bill.bill_date,
        db.func.sum(BillItem.net_amount).label('revenue'),
        db.func.sum(BillItem.quantity).label('units')
    ).join(Bill, Bill.id == BillItem.bill_id)\
     .join(Product, Product.id == BillItem.product_id)\
     .filter(Bill.bill_date >= start_date, Bill.bill_date <= end_date)
    
    if level != 'category' and parent:
        weekly_data = weekly_data.filter(parent_filter)
    
    weekly_results = weekly_data.filter(group_col != '', group_col.isnot(None))\
        .group_by(group_col, Bill.bill_date).order_by(group_col, Bill.bill_date).all()
    
    # Group weekly data by name
    weekly_by_name = {}
    for r in weekly_results:
        name = r.name
        if name not in weekly_by_name:
            weekly_by_name[name] = []
        weekly_by_name[name].append({
            'date': str(r.bill_date),
            'revenue': round(r.revenue or 0, 2),
            'units': r.units or 0
        })
    
    # Build response
    items = []
    for r in current_results:
        name = r.name
        prev_rev = prev_map.get(name, 0)
        curr_rev = r.revenue or 0
        
        if prev_rev > 0:
            growth = ((curr_rev - prev_rev) / prev_rev) * 100
        else:
            growth = 100 if curr_rev > 0 else 0
        
        # Get weekly trend for this item
        weekly = weekly_by_name.get(name, [])
        
        # Calculate trend direction using linear regression
        if len(weekly) >= 2:
            dates_sorted = sorted(set(w['date'] for w in weekly))
            if len(dates_sorted) >= 2:
                rev_by_date = {}
                for w in weekly:
                    rev_by_date[w['date']] = rev_by_date.get(w['date'], 0) + w['revenue']
                X = np.array(range(len(dates_sorted))).reshape(-1, 1)
                y = np.array([rev_by_date[d] for d in dates_sorted])
                model = LinearRegression()
                model.fit(X, y)
                trend_pct = (model.coef_[0] / max(np.mean(y), 0.01)) * 100
            else:
                trend_pct = 0
        else:
            trend_pct = 0
        
        # Status
        if growth > 10:
            status = 'growing'
        elif growth < -10:
            status = 'declining'
        else:
            status = 'stable'
        
        items.append({
            'name': name,
            'revenue': round(curr_rev, 2),
            'units': r.units or 0,
            'bills': r.bills or 0,
            'discount': round(r.discount or 0, 2),
            'growth': round(growth, 2),
            'trend_pct': round(trend_pct, 2),
            'status': status,
            'weekly_trend': weekly,
            'data_points': len(weekly)
        })
    
    # Total stats
    total_revenue = sum(i['revenue'] for i in items)
    total_units = sum(i['units'] for i in items)
    
    # Get top-level options for filtering
    if level == 'category':
        # Return available subcategories for each category
        categories_with_sub = {}
        for item in items:
            cat_name = item['name']
            subcats = db.session.query(Product.section)\
                .filter(Product.division == cat_name, Product.section != '', Product.section.isnot(None))\
                .distinct().all()
            categories_with_sub[cat_name] = [s[0] for s in subcats if s[0]]
    else:
        categories_with_sub = {}
    
    return jsonify({
        'items': items,
        'total_revenue': round(total_revenue, 2),
        'total_units': total_units,
        'total_items': len(items),
        'period': {
            'start': str(start_date),
            'end': str(end_date),
            'weeks': weeks
        },
        'level': level,
        'parent': parent,
        'subcategories': categories_with_sub
    })

@app.route('/api/trends/options', methods=['GET'])
def trends_options():
    """Get available categories, subcategories for filter dropdowns."""
    categories = db.session.query(Product.division)\
        .filter(Product.division != '', Product.division.isnot(None))\
        .distinct().order_by(Product.division).all()
    
    subcategories = db.session.query(Product.section)\
        .filter(Product.section != '', Product.section.isnot(None))\
        .distinct().order_by(Product.section).all()
    
    return jsonify({
        'categories': [c[0] for c in categories if c[0]],
        'subcategories': [s[0] for s in subcategories if s[0]]
    })

@app.route('/api/uploads', methods=['GET'])
def uploads():
    uploads_list = Upload.query.order_by(Upload.uploaded_at.desc()).limit(20).all()
    return jsonify({
        'uploads': [{
            'id': u.id,
            'file_name': u.file_name,
            'file_type': u.file_type,
            'total_rows': u.total_rows,
            'processed_rows': u.processed_rows,
            'status': u.status,
            'uploaded_at': str(u.uploaded_at)
        } for u in uploads_list]
    })

@app.route('/api/data/status', methods=['GET'])
def data_status():
    brands_count = Brand.query.count()
    products_count = Product.query.count()
    bills_count = Bill.query.count()
    items_count = BillItem.query.count()
    uploads_count = Upload.query.count()
    
    # Get date range
    first_bill = Bill.query.order_by(Bill.bill_date).first()
    last_bill = Bill.query.order_by(Bill.bill_date.desc()).first()
    
    return jsonify({
        'brands': brands_count,
        'products': products_count,
        'bills': bills_count,
        'items': items_count,
        'uploads': uploads_count,
        'date_range': {
            'first': str(first_bill.bill_date) if first_bill else None,
            'last': str(last_bill.bill_date) if last_bill else None
        }
    })

@app.route('/api/reset', methods=['POST'])
def reset_data():
    """Clear all data from database."""
    try:
        db.session.query(BillItem).delete()
        db.session.query(Bill).delete()
        db.session.query(Product).delete()
        db.session.query(Brand).delete()
        db.session.query(Upload).delete()
        db.session.commit()
        return jsonify({'success': True, 'message': 'All data cleared'})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/reorder/suggestions', methods=['GET'])
def reorder_suggestions():
    try:
        suggestions = get_reorder_suggestions()
        
        # Group by brand
        brands = {}
        for s in suggestions:
            brand = s['brand']
            if brand not in brands:
                brands[brand] = {'brand': brand, 'items': [], 'total_qty': 0}
            brands[brand]['items'].append(s)
            brands[brand]['total_qty'] += s['reorder_qty']
        
        return jsonify({
            'brands': list(brands.values()),
            'total_brands': len(brands),
            'total_items': len(suggestions),
            'total_reorder_qty': sum(s['reorder_qty'] for s in suggestions)
        })
    except Exception as e:
        import traceback
        print(f"Error in reorder suggestions: {e}")
        print(traceback.format_exc())
        return jsonify({
            'error': str(e),
            'brands': [],
            'total_brands': 0,
            'total_items': 0,
            'total_reorder_qty': 0
        }), 500

@app.route('/api/reorder/export', methods=['POST'])
def reorder_export():
    """Generate PO Excel files per brand."""
    data = request.get_json()
    brand_name = data.get('brand') if data else None
    
    suggestions = get_reorder_suggestions()
    
    # Group by brand
    brands_data = {}
    for s in suggestions:
        brand = s['brand']
        if brand not in brands_data:
            brands_data[brand] = []
        brands_data[brand].append(s)
    
    if brand_name:
        # Export only one brand
        if brand_name not in brands_data:
            return jsonify({'error': f'No suggestions for brand: {brand_name}'}), 404
        
        output_path = f"PO_{brand_name.replace(' ', '_')}.xlsx"
        count = generate_po_for_brand(brand_name, brands_data[brand_name], output_path)
        
        return jsonify({
            'success': True,
            'message': f'Generated PO for {brand_name} with {count} items',
            'file': output_path,
            'brand': brand_name
        })
    else:
        # Export all brands
        files = []
        for brand, items in brands_data.items():
            output_path = f"PO_{brand.replace(' ', '_')}.xlsx"
            count = generate_po_for_brand(brand, items, output_path)
            files.append({
                'brand': brand,
                'items': count,
                'file': output_path
            })
        
        return jsonify({
            'success': True,
            'message': f'Generated {len(files)} PO files',
            'files': files
        })

@app.route('/api/reorder/<path:filename>', methods=['GET'])
def download_po(filename):
    return send_file(filename, as_attachment=True, download_name=filename)

# ─── NEW FEATURES: COMPARE, SUMMARY, REORDER HISTORY ────────────────────────

@app.route('/api/compare', methods=['GET'])
def compare_periods():
    """Compare two periods side-by-side."""
    p1_start = parse_date(request.args.get('p1_start', ''))
    p1_end = parse_date(request.args.get('p1_end', ''))
    p2_start = parse_date(request.args.get('p2_start', ''))
    p2_end = parse_date(request.args.get('p2_end', ''))
    
    if not all([p1_start, p1_end, p2_start, p2_end]):
        # Default: compare last 2 weeks
        today = date.today()
        p2_end = today
        p2_start = today - timedelta(days=6)
        p1_end = p2_start - timedelta(days=1)
        p1_start = p1_end - timedelta(days=6)
    
    def get_period_data(start, end):
        data = get_weekly_data(start, end)
        brands = get_brand_performance(start, end)
        products = get_top_products(start, end, limit=10)
        return {
            'kpis': {
                'revenue': data['total_revenue'],
                'bills': data['total_bills'],
                'items': data['total_items'],
                'discount': data['total_discount'],
                'avg_bill': data['avg_bill_value']
            },
            'brands': brands[:10],
            'products': products,
            'period': {'start': str(start), 'end': str(end)}
        }
    
    p1 = get_period_data(p1_start, p1_end)
    p2 = get_period_data(p2_start, p2_end)
    
    # Calculate changes
    changes = {}
    for key in ['revenue', 'bills', 'items', 'discount', 'avg_bill']:
        v1 = p1['kpis'][key]
        v2 = p2['kpis'][key]
        if v1 > 0:
            changes[key] = round(((v2 - v1) / v1) * 100, 2)
        else:
            changes[key] = 100 if v2 > 0 else 0
    
    return jsonify({
        'period1': p1,
        'period2': p2,
        'changes': changes
    })

@app.route('/api/summary', methods=['GET'])
def business_summary():
    """Generate natural language business summary."""
    week_start_str = request.args.get('week_start')
    week_end_str = request.args.get('week_end')
    
    today = date.today()
    week_start = parse_date(week_start_str) if week_start_str else (today - timedelta(days=today.weekday()))
    week_end = parse_date(week_end_str) if week_end_str else (week_start + timedelta(days=6))
    prev_start, prev_end = get_previous_week_range(week_start)
    
    current = get_weekly_data(week_start, week_end)
    previous = get_weekly_data(prev_start, prev_end)
    brands = get_brand_performance(week_start, week_end)
    prev_brands = get_brand_performance(prev_start, prev_end)
    products = get_top_products(week_start, week_end, limit=5)
    
    # Calculate growth
    rev_growth = 0
    if previous['total_revenue'] > 0:
        rev_growth = ((current['total_revenue'] - previous['total_revenue']) / previous['total_revenue']) * 100
    
    # Build summary sentences
    sentences = []
    
    # Revenue summary
    rev_word = "increased" if rev_growth > 0 else "decreased"
    sentences.append(f"Total revenue was ₹{current['total_revenue']:,.0f}, which {rev_word} by {abs(rev_growth):.1f}% compared to the previous week.")
    
    # Bills summary
    bill_change = current['total_bills'] - previous['total_bills']
    bill_word = "more" if bill_change > 0 else "fewer"
    sentences.append(f"There were {current['total_bills']} bills processed ({abs(bill_change)} {bill_word} than last week), with an average bill value of ₹{current['avg_bill_value']:,.0f}.")
    
    # Top brand
    if brands:
        top = brands[0]
        sentences.append(f"The top performing brand was {top['name']} with ₹{top['revenue']:,.0f} in revenue ({top['share']:.1f}% of total sales).")
    
    # Growing/declining brands
    growing = [b for b in brands if b.get('growth', 0) > 10]
    declining = [b for b in brands if b.get('growth', 0) < -10]
    if growing:
        names = ', '.join(b['name'] for b in growing[:3])
        sentences.append(f"Strong growth from: {names}.")
    if declining:
        names = ', '.join(b['name'] for b in declining[:3])
        sentences.append(f"Declining performance from: {names}.")
    
    # Top product
    if products:
        top_p = products[0]
        sentences.append(f"The best selling product was {top_p['item_name'][:40]} with {top_p['units']} units sold (₹{top_p['revenue']:,.0f}).")
    
    # Discount summary
    if current['total_discount'] > 0:
        disc_pct = (current['total_discount'] / current['total_revenue']) * 100 if current['total_revenue'] > 0 else 0
        sentences.append(f"Total discounts given were ₹{current['total_discount']:,.0f} ({disc_pct:.1f}% of revenue).")
    
    return jsonify({
        'summary': ' '.join(sentences),
        'sentences': sentences,
        'period': {'start': str(week_start), 'end': str(week_end)},
        'highlights': {
            'total_revenue': current['total_revenue'],
            'revenue_growth': round(rev_growth, 2),
            'total_bills': current['total_bills'],
            'top_brand': brands[0]['name'] if brands else None,
            'top_brand_revenue': brands[0]['revenue'] if brands else 0,
            'growing_brands': len(growing),
            'declining_brands': len(declining)
        }
    })

@app.route('/api/reorder/history', methods=['GET'])
def reorder_history():
    """Get reorder history."""
    history = ReorderHistory.query.order_by(ReorderHistory.order_date.desc()).limit(50).all()
    return jsonify({
        'history': [{
            'id': h.id,
            'brand': h.brand_name,
            'barcode': h.barcode,
            'product': h.product_name,
            'qty_ordered': h.qty_ordered,
            'qty_sold_since': h.qty_sold_since,
            'order_date': str(h.order_date.date()) if h.order_date else '',
            'status': h.status,
            'notes': h.notes
        } for h in history]
    })

@app.route('/api/reorder/history', methods=['POST'])
def save_reorder_history():
    """Save reorder history when PO is exported."""
    data = request.get_json()
    items = data.get('items', [])
    count = 0
    for item in items:
        # Calculate how many units sold since order
        order_date = datetime.utcnow()
        start_date = order_date - timedelta(days=30)
        qty_sold = db.session.query(db.func.sum(BillItem.quantity)).join(Bill)\
            .filter(BillItem.barcode == item.get('barcode', ''))\
            .filter(Bill.bill_date >= start_date.date()).scalar() or 0
        
        record = ReorderHistory(
            brand_name=item.get('brand', ''),
            barcode=item.get('barcode', ''),
            product_name=item.get('product_name', ''),
            qty_ordered=item.get('qty', 0),
            qty_sold_since=qty_sold,
            status='ordered'
        )
        db.session.add(record)
        count += 1
    db.session.commit()
    return jsonify({'success': True, 'message': f'Saved {count} reorder records'})

@app.route('/api/filters/options', methods=['GET'])
def filter_options():
    """Get all filter options for interactive filtering."""
    brands = [b.name for b in Brand.query.order_by(Brand.name).all()]
    categories = db.session.query(Product.division).filter(Product.division != '', Product.division.isnot(None)).distinct().order_by(Product.division).all()
    departments = db.session.query(Product.department).filter(Product.department != '', Product.department.isnot(None)).distinct().order_by(Product.department).all()
    
    return jsonify({
        'brands': brands,
        'categories': [c[0] for c in categories if c[0]],
        'departments': [d[0] for d in departments if d[0]]
    })

# ─── PRODUCT BREAKOUT PREDICTION ────────────────────────────────────────────

@app.route('/api/predictions/breakout', methods=['GET'])
def product_breakout_prediction():
    """Detect products likely to suddenly become best sellers."""
    # Get top products by revenue for current and previous periods
    today = date.today()
    current_end = today - timedelta(days=today.weekday())
    current_start = current_end - timedelta(weeks=2)
    prev_end = current_start - timedelta(days=1)
    prev_start = prev_end - timedelta(weeks=2)
    
    # Current period rankings
    current_products = get_top_products(current_start, current_end, limit=100)
    prev_products = get_top_products(prev_start, prev_end, limit=100)
    
    # Create ranking maps
    current_rank_map = {p['barcode']: i+1 for i, p in enumerate(current_products)}
    prev_rank_map = {p['barcode']: i+1 for i, p in enumerate(prev_products)}
    
    # Get all products with recent sales
    all_barcodes = set(list(current_rank_map.keys()) + list(prev_rank_map.keys()))
    
    breakout_candidates = []
    
    for barcode in all_barcodes:
        # Get product details
        product = Product.query.filter_by(barcode=barcode).first()
        if not product:
            continue
        
        current_rank = current_rank_map.get(barcode, 999)
        prev_rank = prev_rank_map.get(barcode, 999)
        
        # Get current period sales
        current_sales = next((p for p in current_products if p['barcode'] == barcode), None)
        prev_sales = next((p for p in prev_products if p['barcode'] == barcode), None)
        
        current_revenue = current_sales['revenue'] if current_sales else 0
        prev_revenue = prev_sales['revenue'] if prev_sales else 0
        
        # Calculate acceleration
        if prev_revenue > 0:
            acceleration = ((current_revenue - prev_revenue) / prev_revenue) * 100
        else:
            acceleration = 100 if current_revenue > 0 else 0
        
        # Calculate rank change
        rank_change = prev_rank - current_rank
        
        # Only consider products that were in BOTH periods and are improving
        if acceleration <= 0 or rank_change <= 0 or prev_rank >= 999:
            continue
        
        # Calculate confidence score
        confidence = 50  # Base confidence
        
        # Boost confidence based on:
        # 1. Strong acceleration (>50%)
        if acceleration > 50:
            confidence += 20
        elif acceleration > 30:
            confidence += 15
        elif acceleration > 15:
            confidence += 10
        
        # 2. Significant rank improvement
        if rank_change >= 10:
            confidence += 20
        elif rank_change >= 5:
            confidence += 15
        elif rank_change >= 3:
            confidence += 10
        
        # 3. Data consistency (both periods have data)
        if current_sales and prev_sales:
            confidence += 10
        
        # Cap at 95%
        confidence = min(confidence, 95)
        
        # Determine classification
        if rank_change >= 5 and acceleration > 30:
            classification = 'BREAKOUT'
            emoji = '🔥'
        elif rank_change >= 2 and acceleration > 15:
            classification = 'RISING'
            emoji = '📈'
        else:
            classification = 'STABLE'
            emoji = '➡️'
        
        # Generate reason
        if acceleration > 50:
            reason = f"Sales accelerating +{acceleration:.0f}%"
        elif rank_change >= 10:
            reason = f"Rank jumped {rank_change} positions"
        else:
            reason = f"Steady growth +{acceleration:.0f}%"
        
        breakout_candidates.append({
            'barcode': barcode,
            'product_name': product.item_name or 'Unknown',
            'brand': product.brand.name if product.brand else 'Unknown',
            'current_rank': current_rank,
            'predicted_rank': max(1, current_rank - int(rank_change * 0.8)),
            'rank_change': rank_change,
            'current_revenue': round(current_revenue, 2),
            'acceleration': round(acceleration, 1),
            'confidence': confidence,
            'classification': classification,
            'emoji': emoji,
            'reason': reason
        })
    
    # Sort by confidence and rank change
    breakout_candidates.sort(key=lambda x: (x['confidence'], x['rank_change']), reverse=True)
    
    # Save predictions to database
    for candidate in breakout_candidates[:20]:  # Save top 20
        pred = Prediction(
            prediction_type='product_breakout',
            predicted_value=candidate['predicted_rank'],
            predicted_date=today + timedelta(weeks=1),
            confidence=candidate['confidence'],
            details=json.dumps({
                'barcode': candidate['barcode'],
                'product_name': candidate['product_name'],
                'brand': candidate['brand'],
                'current_rank': candidate['current_rank'],
                'rank_change': candidate['rank_change'],
                'acceleration': candidate['acceleration'],
                'classification': candidate['classification']
            })
        )
        db.session.add(pred)
    
    db.session.commit()
    
    return jsonify({
        'breakouts': breakout_candidates[:20],  # Top 20
        'total_candidates': len(breakout_candidates),
        'period': {
            'current': f"{current_start} to {current_end}",
            'previous': f"{prev_start} to {prev_end}"
        }
    })

# ─── PREDICTION ACCURACY TRACKER ────────────────────────────────────────────

@app.route('/api/predictions/accuracy', methods=['GET'])
def prediction_accuracy():
    """Track prediction accuracy over time."""
    # Get revenue predictions
    revenue_predictions = Prediction.query.filter_by(prediction_type='revenue')\
        .order_by(Prediction.created_at.desc()).limit(20).all()
    
    # Calculate accuracy metrics
    total_predictions = len(revenue_predictions)
    accurate_predictions = 0
    total_error = 0
    
    accuracy_history = []
    
    for pred in revenue_predictions:
        if pred.actual_value is not None and pred.predicted_value is not None:
            error = abs(pred.predicted_value - pred.actual_value)
            error_pct = (error / pred.actual_value * 100) if pred.actual_value > 0 else 0
            accuracy = max(0, 100 - error_pct)
            
            total_error += error_pct
            if error_pct < 10:  # Consider <10% error as accurate
                accurate_predictions += 1
            
            accuracy_history.append({
                'date': str(pred.predicted_date),
                'predicted': pred.predicted_value,
                'actual': pred.actual_value,
                'error': round(error_pct, 2),
                'accuracy': round(accuracy, 2)
            })
    
    avg_accuracy = (accurate_predictions / total_predictions * 100) if total_predictions > 0 else 0
    avg_error = (total_error / total_predictions) if total_predictions > 0 else 0
    
    # Get latest prediction
    latest = revenue_predictions[0] if revenue_predictions else None
    
    return jsonify({
        'summary': {
            'total_predictions': total_predictions,
            'avg_accuracy': round(avg_accuracy, 2),
            'avg_error': round(avg_error, 2),
            'accurate_count': accurate_predictions
        },
        'latest': {
            'predicted': latest.predicted_value if latest else 0,
            'actual': latest.actual_value if latest else 0,
            'error': round(abs(latest.predicted_value - latest.actual_value) / latest.actual_value * 100, 2) if latest and latest.actual_value and latest.actual_value > 0 else 0,
            'accuracy': round(latest.accuracy_percent, 2) if latest else 0,
            'date': str(latest.predicted_date) if latest else None
        } if latest else None,
        'history': accuracy_history[:10]  # Last 10 predictions
    })

@app.route('/api/predictions/verify', methods=['POST'])
def verify_prediction():
    """Verify a prediction with actual data."""
    data = request.get_json()
    prediction_id = data.get('prediction_id')
    actual_value = data.get('actual_value')
    
    pred = Prediction.query.get(prediction_id)
    if not pred:
        return jsonify({'error': 'Prediction not found'}), 404
    
    pred.actual_value = actual_value
    pred.actual_date = date.today()
    
    # Calculate error and accuracy
    if pred.predicted_value and pred.actual_value:
        error = abs(pred.predicted_value - pred.actual_value)
        pred.error_percent = (error / pred.actual_value * 100) if pred.actual_value > 0 else 0
        pred.accuracy_percent = max(0, 100 - pred.error_percent)
    
    pred.verified_at = datetime.utcnow()
    db.session.commit()
    
    return jsonify({
        'success': True,
        'error_percent': pred.error_percent,
        'accuracy_percent': pred.accuracy_percent
    })

# ─── AI MERCHANDISING ───────────────────────────────────────────────────────

@app.route('/api/ai_merchandising', methods=['GET'])
def ai_merchandising():
    """Generates seasonal forecasts and styling combinations using Gemini AI."""
    try:
        force = request.args.get('force', 'false').lower() == 'true'
        current_month_str = date.today().strftime("%Y_%m")
        cache_dir = os.path.join(app.config.get('UPLOAD_FOLDER', 'uploads'), 'cache')
        os.makedirs(cache_dir, exist_ok=True)
        cache_file = os.path.join(cache_dir, f"ai_merchandising_cache_{current_month_str}.json")
        
        if not force and os.path.exists(cache_file):
            with open(cache_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                return jsonify({"success": True, "data": data, "cached": True})
                
        # 1. Fetch unique product attributes from database to send to AI
        products = db.session.query(
            Product.category2,
            Product.department,
            Product.color,
            Product.material,
            Product.gender,
            Brand.name.label('brand_name')
        ).join(Brand, Product.brand_id == Brand.id).distinct().limit(150).all()

        inventory_summary = []
        for p in products:
            item_desc = f"{p.brand_name} {p.gender} {p.color} {p.material} {p.category2} (Dept: {p.department})"
            item_desc = item_desc.replace("  ", " ").strip()
            if len(item_desc) > 5 and 'None' not in item_desc:
                inventory_summary.append(item_desc)
        
        # Deduplicate and limit
        inventory_summary = list(set(inventory_summary))[:100]
        inventory_text = "\\n".join(inventory_summary)

        if not inventory_text:
             return jsonify({"success": False, "error": "Not enough inventory data loaded to make AI suggestions."})

        current_month = date.today().strftime("%B")
        
        prompt = f"""
        You are an expert fashion merchandiser, retail strategist, and stylist.
        The current month is {current_month}.
        
        Here is a sample of the products currently available in my store's inventory:
        {inventory_text}
        
        Based ONLY on the inventory provided above, please provide two things in JSON format:
        
        1. "seasonal_forecast": Based on the current month ({current_month}) and upcoming real-world Indian festivals/events in the next 1 month ONLY, what 10 categories or types of items from the inventory should I heavily stock up on or promote? For each, provide a "title" (e.g. "Diwali Ethnic Wear"), "reason" (why this is relevant for the upcoming Indian festival/event), and "items" (a list of 2-3 specific matching item descriptions from my inventory).
        
        2. "stylist_picks": Create 10 stylish product combinations (outfits or bundles) that look great together from the inventory. For each, provide a "title" (e.g. "Summer Casual Combo"), "reason" (why they match in color/style), and "items" (a list of 2-4 specific item descriptions from my inventory that make up the combo).
        
        Return ONLY valid JSON with no markdown formatting or extra text. The JSON structure should be:
        {{
            "seasonal_forecast": [
                {{"title": "...", "reason": "...", "items": ["...", "..."]}}
            ],
            "stylist_picks": [
                {{"title": "...", "reason": "...", "items": ["...", "..."]}}
            ]
        }}
        """

        response = co.chat(
            model="command-r-plus-08-2024",
            message=prompt,
            temperature=0.3
        )
        
        response_text = response.text.strip()
        if response_text.startswith("```json"):
            response_text = response_text[7:]
        if response_text.endswith("```"):
            response_text = response_text[:-3]
            
        data = json.loads(response_text)
        
        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(data, f)
            
        return jsonify({"success": True, "data": data, "cached": False})

    except Exception as e:
        print(f"AI Merchandising error: {e}")
        return jsonify({"success": False, "error": str(e)}), 500

# ─── BUYING PLAN INTEGRATION ────────────────────────────────────────────────

from buying_plan_loader import load_buying_plan, get_buying_plan_products, get_ai_buying_suggestions, clear_cache

@app.route('/api/buying-plan/load', methods=['GET'])
def buying_plan_load():
    """Load buying plan Excel data."""
    force = request.args.get('force', 'false').lower() == 'true'
    if force:
        clear_cache()
    data = load_buying_plan()
    return jsonify(data)

@app.route('/api/buying-plan/products', methods=['GET'])
def buying_plan_products():
    """Get filtered buying plan products."""
    brand = request.args.get('brand')
    category = request.args.get('category')
    sub_category = request.args.get('sub_category')
    search = request.args.get('search')
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 50, type=int)
    
    products = get_buying_plan_products(brand, category, sub_category, search)
    total = len(products)
    
    # Paginate
    start = (page - 1) * per_page
    end = start + per_page
    page_products = products[start:end]
    
    return jsonify({
        'products': page_products,
        'total': total,
        'page': page,
        'per_page': per_page,
        'total_pages': (total + per_page - 1) // per_page
    })

@app.route('/api/buying-plan/ai-suggestions', methods=['GET'])
def buying_plan_ai_suggestions():
    """Get AI-driven suggestions from buying plan based on POS sales data."""
    limit = request.args.get('limit', 20, type=int)
    try:
        suggestions = get_ai_buying_suggestions(limit=limit, db=db)
        return jsonify(suggestions)
    except Exception as e:
        import traceback
        print(f"Error in AI suggestions: {e}")
        print(traceback.format_exc())
        return jsonify({
            'suggestions': [],
            'total': 0,
            'error': str(e)
        })

@app.route('/api/buying-plan/image/<brand>/<filename>')
def buying_plan_image(brand, filename):
    """Serve product images from buying plan app's product_images directory."""
    from pathlib import Path
    img_path = Path(__file__).parent.parent / "buying_plan_app" / "product_images" / brand / filename
    if img_path.exists():
        return send_file(str(img_path))
    return jsonify({'error': 'Image not found'}), 404

@app.route('/api/buying-plan/add-to-reorder', methods=['POST'])
def buying_plan_add_to_reorder():
    """Add selected buying plan items to the reorder suggestions list (stored in session)."""
    data = request.get_json()
    items = data.get('items', []) if data else []
    
    if 'buying_plan_selected' not in session:
        session['buying_plan_selected'] = []
    
    existing = session['buying_plan_selected']
    existing_barcodes = set(item.get('barcode') for item in existing)
    
    added = 0
    for item in items:
        barcode = item.get('barcode', '')
        if barcode and barcode not in existing_barcodes:
            existing.append({
                'barcode': barcode,
                'brand': item.get('brand', ''),
                'product_name': item.get('product_title', ''),
                'category': item.get('category', ''),
                'sub_category': item.get('sub_category', ''),
                'mrp': item.get('mrp'),
                'color': item.get('color', ''),
                'size': item.get('size', ''),
                'material': item.get('material', ''),
                'gender': item.get('gender', ''),
                'image_url': item.get('image_url', '') or item.get('local_image', '') or item.get('display_image', ''),
                'reorder_qty': item.get('quantity', 4),
                'source': 'buying_plan'
            })
            existing_barcodes.add(barcode)
            added += 1
    
    session['buying_plan_selected'] = existing
    session.modified = True
    
    return jsonify({
        'success': True,
        'added': added,
        'total_selected': len(existing)
    })

@app.route('/api/buying-plan/selected', methods=['GET'])
def buying_plan_selected():
    """Get items selected from buying plan (stored in session)."""
    items = session.get('buying_plan_selected', [])
    return jsonify({
        'items': items,
        'total': len(items)
    })

@app.route('/api/buying-plan/clear-selected', methods=['POST'])
def buying_plan_clear_selected():
    """Clear selected buying plan items from session."""
    session['buying_plan_selected'] = []
    session.modified = True
    return jsonify({'success': True})

if __name__ == '__main__':
    app.run(debug=True, port=5000)
